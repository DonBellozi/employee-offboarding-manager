from __future__ import annotations

import io
import json
import re
import base64
import hashlib
from dataclasses import dataclass, field
from datetime import datetime, timezone

from cryptography.fernet import Fernet, InvalidToken
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.config import Settings
from app.models import AuditLog, EmailLoginMapping, HRSourceRecord
from app.models_onec_sources import HREmploymentState
from app.models_techexpert import (
    TechExpertActualizationFile,
    TechExpertActualizationItem,
    TechExpertActualizationRun,
    TechExpertSettings,
)
from app.services.ad import ADDirectoryUser, ActiveDirectoryService
from app.services.hr_registry_manual_mapping import (
    HRRegistryManualMappingService,
)
from app.services.techexpert_access import (
    normalize_email,
    normalize_fio,
    normalize_text,
    placement_snapshot,
)


ACTIVE_EMPLOYMENT_STATUSES = {"active", "scheduled"}
MAX_FILE_SIZE = 10 * 1024 * 1024
MAX_DATA_ROWS = 10_000
COMPATIBLE_CORPORATE_TLDS = {"com", "ru"}


def utcnow() -> datetime:
    return datetime.now(timezone.utc)


def normalize_header(value: object) -> str:
    return re.sub(r"[^0-9a-zа-я]+", "", normalize_text(value).casefold())


def safe_excel_value(value: object) -> str:
    text = normalize_text(value)
    if text.startswith(("=", "+", "-", "@")):
        return "'" + text
    return text


def safe_excel_secret_value(value: object) -> str:
    """Сохранить пароль без нормализации, но исключить формулы Excel."""

    text = "" if value is None else str(value)
    if text.startswith(("=", "+", "-", "@")):
        return "'" + text
    return text


def organization_email_keys(value: object, source_id: object) -> set[str]:
    """Exact email plus a safe .com/.ru alias for the configured organization."""

    email = normalize_email(value)
    current_domain = normalize_email(source_id).lstrip("@")
    if email.count("@") != 1 or not current_domain:
        return {email} if email else set()

    local_part, domain = email.rsplit("@", 1)
    current_stem, current_dot, current_tld = current_domain.rpartition(".")
    domain_stem, domain_dot, domain_tld = domain.rpartition(".")
    if (
        local_part
        and current_dot
        and domain_dot
        and current_stem == domain_stem
        and current_tld in COMPATIBLE_CORPORATE_TLDS
        and domain_tld in COMPATIBLE_CORPORATE_TLDS
    ):
        return {email, f"{local_part}@{current_domain}"}
    return {email}


@dataclass(frozen=True)
class ParsedTechExpertRow:
    row_number: int
    fio: str
    position: str
    email: str
    phone: str
    login: str
    password: str = field(repr=False)


@dataclass(frozen=True)
class ParsedTechExpertFile:
    department_name: str
    rows: tuple[ParsedTechExpertRow, ...]


@dataclass(frozen=True)
class ActualizationAnalysisContext:
    active_worker_keys: set[str]
    by_fio: dict[str, list[HRSourceRecord]]
    by_email: dict[str, list[HRSourceRecord]]
    mappings: dict[str, list[EmailLoginMapping]]
    ad: ActiveDirectoryService
    member_by_guid: dict[str, ADDirectoryUser]
    member_by_login: dict[str, ADDirectoryUser]
    group_error: str


class TechExpertCredentialBox:
    """Шифрует исходные пароли ТЭ ключом приложения."""

    def __init__(self, app_secret_key: str):
        secret = str(app_secret_key or "").encode("utf-8")
        if not secret:
            raise RuntimeError("APP_SECRET_KEY не задан")
        digest = hashlib.sha256(secret).digest()
        self._fernet = Fernet(base64.urlsafe_b64encode(digest))

    def encrypt(self, value: str) -> str:
        text = str(value or "")
        if not text:
            return ""
        return self._fernet.encrypt(text.encode("utf-8")).decode("ascii")

    def decrypt(self, value: str) -> str:
        text = str(value or "")
        if not text:
            return ""
        try:
            return self._fernet.decrypt(text.encode("ascii")).decode("utf-8")
        except (InvalidToken, ValueError, UnicodeError) as exc:
            raise RuntimeError(
                "Не удалось расшифровать пароль Техэксперта. "
                "Проверьте, что APP_SECRET_KEY не менялся."
            ) from exc


class TechExpertActualizationService:
    """Первичная и редкая сверка файлов с кадровой организацией и AD."""

    def __init__(
        self,
        settings: Settings,
        db: Session,
        config: TechExpertSettings,
    ):
        self.settings = settings
        self.db = db
        self.config = config
        self.credential_box = TechExpertCredentialBox(settings.app_secret_key)

    @property
    def source_id(self) -> str:
        return normalize_email(self.config.source_domain)

    def _require_configuration(self) -> None:
        if not self.source_id:
            raise ValueError("Сначала выберите организацию Техэксперта")
        if not str(self.config.ad_group_dn or "").strip():
            raise ValueError("Сначала сохраните DN группы AD Техэксперта")

    @staticmethod
    def parse_xlsx(data: bytes) -> ParsedTechExpertFile:
        if not data:
            raise ValueError("Файл пуст")
        if len(data) > MAX_FILE_SIZE:
            raise ValueError("Файл больше 10 МБ")

        workbook = load_workbook(
            io.BytesIO(data),
            read_only=True,
            data_only=True,
            keep_links=False,
        )
        try:
            sheet = workbook.active
            department_name = next(
                (
                    normalize_text(sheet.cell(3, column).value)
                    for column in range(1, sheet.max_column + 1)
                    if normalize_text(sheet.cell(3, column).value)
                ),
                "",
            )
            if not department_name:
                raise ValueError("В строке 3 не найдено название подразделения")

            aliases = {
                "fio": {"фио", "фамилияимяотчество"},
                "position": {"должность"},
                "email": {"email", "электроннаяпочта", "почта"},
                "phone": {"телефон", "тел"},
                "login": {"логин", "username", "имяпользователя"},
                "password": {"пароль", "password"},
            }
            columns: dict[str, int] = {}
            for column in range(1, sheet.max_column + 1):
                header = normalize_header(sheet.cell(4, column).value)
                for field, field_aliases in aliases.items():
                    if header in field_aliases and field not in columns:
                        columns[field] = column
            if "fio" not in columns:
                raise ValueError("В строке 4 не найдена колонка «ФИО»")

            rows: list[ParsedTechExpertRow] = []
            last_row = min(sheet.max_row, 4 + MAX_DATA_ROWS)
            for row_number in range(5, last_row + 1):
                fio = normalize_text(
                    sheet.cell(row_number, columns["fio"]).value
                )
                if not fio:
                    continue
                rows.append(
                    ParsedTechExpertRow(
                        row_number=row_number,
                        fio=fio,
                        position=normalize_text(
                            sheet.cell(
                                row_number,
                                columns.get("position", 0),
                            ).value
                            if columns.get("position")
                            else ""
                        ),
                        email=normalize_email(
                            sheet.cell(
                                row_number,
                                columns.get("email", 0),
                            ).value
                            if columns.get("email")
                            else ""
                        ),
                        phone=normalize_text(
                            sheet.cell(
                                row_number,
                                columns.get("phone", 0),
                            ).value
                            if columns.get("phone")
                            else ""
                        ),
                        login=normalize_text(
                            sheet.cell(
                                row_number,
                                columns.get("login", 0),
                            ).value
                            if columns.get("login")
                            else ""
                        ),
                        password=(
                            ""
                            if not columns.get("password")
                            or sheet.cell(
                                row_number,
                                columns["password"],
                            ).value is None
                            else str(
                                sheet.cell(
                                    row_number,
                                    columns["password"],
                                ).value
                            )
                        ),
                    )
                )
            if sheet.max_row > last_row:
                raise ValueError(
                    f"В файле больше допустимых {MAX_DATA_ROWS} строк данных"
                )
            if not rows:
                raise ValueError("В файле не найдено ни одного сотрудника")
            return ParsedTechExpertFile(department_name, tuple(rows))
        finally:
            workbook.close()

    def create_run(self, *, actor: str, title: str = "") -> TechExpertActualizationRun:
        self._require_configuration()
        existing = self.db.scalar(
            select(TechExpertActualizationRun)
            .where(
                TechExpertActualizationRun.source_id == self.source_id,
                TechExpertActualizationRun.status == "open",
            )
            .order_by(TechExpertActualizationRun.id.desc())
        )
        if existing is not None:
            return existing
        run = TechExpertActualizationRun(
            source_id=self.source_id,
            title=normalize_text(title) or "Первичная сверка Техэксперта",
            created_by=actor,
        )
        self.db.add(run)
        self.db.commit()
        self.db.refresh(run)
        return run

    def get_run(self, run_id: int) -> TechExpertActualizationRun:
        run = self.db.get(TechExpertActualizationRun, int(run_id))
        if run is None or normalize_email(run.source_id) != self.source_id:
            raise LookupError("Пакет актуализации не найден")
        return run

    def _source_records(self) -> list[HRSourceRecord]:
        return list(
            self.db.scalars(
                select(HRSourceRecord).where(
                    HRSourceRecord.source_id == self.source_id
                )
            ).all()
        )

    def _active_worker_keys(self) -> set[str]:
        return {
            str(state.worker_key or "").strip()
            for state in self.db.scalars(
                select(HREmploymentState).where(
                    HREmploymentState.source_id == self.source_id,
                    HREmploymentState.status.in_(ACTIVE_EMPLOYMENT_STATUSES),
                )
            ).all()
            if str(state.worker_key or "").strip()
        }

    def _mappings(self, worker_keys: set[str]) -> dict[str, list[EmailLoginMapping]]:
        result: dict[str, list[EmailLoginMapping]] = {}
        if not worker_keys:
            return result
        for mapping in self.db.scalars(
            select(EmailLoginMapping).where(
                EmailLoginMapping.worker_key.in_(worker_keys)
            )
        ).all():
            result.setdefault(mapping.worker_key, []).append(mapping)
        return result

    @staticmethod
    def _candidate_payload(record: HRSourceRecord) -> dict[str, object]:
        snapshot = placement_snapshot(record)
        return {
            "record_id": record.id,
            "worker_key": record.worker_key,
            "fio": record.fio,
            "positions": snapshot["positions"],
            "departments": snapshot["departments"],
            "corporate_email": record.corporate_email,
            "personal_email": record.personal_email,
        }

    def _resolve_ad(
        self,
        record: HRSourceRecord,
        mappings: dict[str, list[EmailLoginMapping]],
        ad: ActiveDirectoryService,
        member_by_guid: dict[str, ADDirectoryUser],
        member_by_login: dict[str, ADDirectoryUser],
    ) -> ADDirectoryUser | None:
        values = mappings.get(record.worker_key, [])
        preferred = next(
            (
                value
                for value in values
                if normalize_email(value.source_domain) == self.source_id
            ),
            None,
        )
        if preferred is None:
            identities = {
                (
                    normalize_email(value.ad_object_guid),
                    normalize_email(value.ad_login),
                )
                for value in values
                if value.ad_object_guid or value.ad_login
            }
            if len(identities) == 1:
                guid, login = next(iter(identities))
            elif len(identities) > 1:
                return None
            else:
                guid, login = "", normalize_email(record.login)
        else:
            guid = normalize_email(preferred.ad_object_guid)
            login = normalize_email(preferred.ad_login)

        user = member_by_guid.get(guid) if guid else None
        if user is None and login:
            user = member_by_login.get(login)
        if user is None and guid:
            user = ad.get_user_by_object_guid(guid)
        if user is None and login:
            user = ad.get_user(login)
        if user is None and record.corporate_email:
            candidates = ad.users_by_email(record.corporate_email, limit=5)
            if len(candidates) == 1:
                user = candidates[0]
        if user is None:
            candidates = [
                value
                for value in ad.search_users(record.fio, limit=10)
                if normalize_fio(value.display_name) == normalize_fio(record.fio)
            ]
            unique = {value.object_guid or value.username: value for value in candidates}
            if len(unique) == 1:
                user = next(iter(unique.values()))
        return user

    def _analysis_context(self) -> ActualizationAnalysisContext:
        records = self._source_records()
        by_fio: dict[str, list[HRSourceRecord]] = {}
        by_email: dict[str, list[HRSourceRecord]] = {}
        for record in records:
            by_fio.setdefault(normalize_fio(record.fio), []).append(record)
            for email in (record.corporate_email, record.personal_email):
                for email_key in organization_email_keys(email, self.source_id):
                    by_email.setdefault(email_key, []).append(record)

        ad = ActiveDirectoryService(self.settings)
        try:
            members = ad.group_members(self.config.ad_group_dn)
            group_error = ""
        except Exception as exc:
            members = []
            group_error = str(exc)
        return ActualizationAnalysisContext(
            active_worker_keys=self._active_worker_keys(),
            by_fio=by_fio,
            by_email=by_email,
            mappings=self._mappings({record.worker_key for record in records}),
            ad=ad,
            member_by_guid={
                normalize_email(member.object_guid): member
                for member in members
                if member.object_guid
            },
            member_by_login={
                normalize_email(member.username): member for member in members
            },
            group_error=group_error,
        )

    def _records_by_email(
        self,
        context: ActualizationAnalysisContext,
        value: object,
    ) -> list[HRSourceRecord]:
        matched: dict[int, HRSourceRecord] = {}
        for email_key in organization_email_keys(value, self.source_id):
            for candidate in context.by_email.get(email_key, []):
                matched[candidate.id] = candidate
        return list(matched.values())

    def _analyze_source(
        self,
        source: ParsedTechExpertRow,
        context: ActualizationAnalysisContext,
    ) -> dict[str, object]:
        exact = list(context.by_fio.get(normalize_fio(source.fio), []))
        unique = {record.worker_key: record for record in exact}
        candidates = list(unique.values())
        reason_parts: list[str] = []

        if len(candidates) > 1 and source.email:
            email_candidates = {
                record.worker_key: record
                for record in self._records_by_email(context, source.email)
                if record.worker_key in unique
            }
            if len(email_candidates) == 1:
                candidates = list(email_candidates.values())
                reason_parts.append(
                    "Совпадение уточнено по e-mail"
                    if source.email in context.by_email
                    else "Совпадение уточнено по корпоративному e-mail "
                    "с учетом смены домена .com/.ru"
                )

        record = candidates[0] if len(candidates) == 1 else None
        if record is None:
            email_suggestions = (
                {
                    value.worker_key: value
                    for value in self._records_by_email(context, source.email)
                }
                if source.email
                else {}
            )
            display_candidates = candidates or list(email_suggestions.values())
            if not exact and not email_suggestions:
                category = "not_working"
                reason_parts.append(
                    "ФИО отсутствует среди работников организации"
                )
            elif not exact and email_suggestions:
                category = "review"
                reason_parts.append("По e-mail найден сотрудник с другим ФИО")
            else:
                category = "review"
                reason_parts.append("Найдено несколько работников с таким ФИО")
            return {
                "category": category,
                "reason": "; ".join(reason_parts),
                "candidates_json": json.dumps(
                    [self._candidate_payload(value) for value in display_candidates],
                    ensure_ascii=False,
                ),
                "worker_key": "",
                "hr_record_id": None,
                "current_fio": "",
                "current_positions": "[]",
                "current_departments": "[]",
                "ad_login": "",
                "ad_object_guid": "",
                "ad_distinguished_name": "",
                "ad_status": "not_found",
                "membership_state": (
                    "error" if context.group_error else "not_checked"
                ),
            }

        snapshot = placement_snapshot(record)
        normalized_positions = {
            normalize_fio(value) for value in snapshot["positions"]
        }
        if (
            source.position
            and normalize_fio(source.position) not in normalized_positions
        ):
            reason_parts.append("Должность в кадрах изменилась")
        record_email_keys = {
            email_key
            for value in (record.corporate_email, record.personal_email)
            for email_key in organization_email_keys(value, self.source_id)
        }
        if (
            source.email
            and not organization_email_keys(source.email, self.source_id)
            & record_email_keys
        ):
            reason_parts.append("E-mail отличается от кадровых данных")

        try:
            ad_user = self._resolve_ad(
                record,
                context.mappings,
                context.ad,
                context.member_by_guid,
                context.member_by_login,
            )
            ad_error = ""
        except Exception as exc:
            ad_user = None
            ad_error = str(exc)

        membership_state = "error" if context.group_error else "not_member"
        if ad_user is not None and (
            normalize_email(ad_user.object_guid) in context.member_by_guid
            or normalize_email(ad_user.username) in context.member_by_login
        ):
            membership_state = "member"

        is_active = record.worker_key in context.active_worker_keys
        if is_active and ad_user is not None and ad_user.is_enabled:
            category = "working"
        elif is_active:
            category = "review"
            reason_parts.append(
                "AD-учетка не найдена"
                if ad_user is None
                else "AD-учетка отключена"
            )
        else:
            category = "not_working"
            reason_parts.append("Нет активной занятости в организации")

        if ad_error:
            reason_parts.append(f"AD: {ad_error}")
        if context.group_error:
            reason_parts.append(f"Группа AD: {context.group_error}")
        return {
            "category": category,
            "reason": "; ".join(dict.fromkeys(reason_parts)),
            "candidates_json": json.dumps(
                [self._candidate_payload(record)],
                ensure_ascii=False,
            ),
            "worker_key": record.worker_key,
            "hr_record_id": record.id,
            "current_fio": record.fio,
            "current_positions": json.dumps(
                snapshot["positions"], ensure_ascii=False
            ),
            "current_departments": json.dumps(
                snapshot["departments"], ensure_ascii=False
            ),
            "ad_login": ad_user.username if ad_user else "",
            "ad_object_guid": ad_user.object_guid if ad_user else "",
            "ad_distinguished_name": (
                ad_user.distinguished_name if ad_user else ""
            ),
            "ad_status": (
                "enabled"
                if ad_user is not None and ad_user.is_enabled
                else "disabled"
                if ad_user is not None
                else "error"
                if ad_error
                else "not_found"
            ),
            "membership_state": membership_state,
        }

    @staticmethod
    def _apply_analysis(
        item: TechExpertActualizationItem,
        analysis: dict[str, object],
    ) -> None:
        previous_category = item.category
        for field, value in analysis.items():
            setattr(item, field, value)
        if item.category != previous_category:
            item.group_action = "not_started"
            item.group_action_error = ""

    def add_file(
        self,
        *,
        run_id: int,
        filename: str,
        data: bytes,
        actor: str,
    ) -> dict[str, int | str]:
        run = self.get_run(run_id)
        if run.status != "open":
            raise ValueError("Пакет уже завершён")
        if not normalize_text(filename).casefold().endswith(".xlsx"):
            raise ValueError("Загрузите файл XLSX")
        parsed = self.parse_xlsx(data)
        display_filename = normalize_text(filename)
        stem, dot, extension = display_filename.rpartition(".")
        if not dot:
            stem, extension = display_filename, "xlsx"
        suffix = 1
        while self.db.scalar(
            select(TechExpertActualizationFile.id).where(
                TechExpertActualizationFile.run_id == run.id,
                TechExpertActualizationFile.filename == display_filename,
            )
        ) is not None:
            suffix += 1
            display_filename = f"{stem} ({suffix}).{extension}"

        context = self._analysis_context()

        file_row = TechExpertActualizationFile(
            run_id=run.id,
            filename=display_filename,
            department_name=parsed.department_name,
        )
        self.db.add(file_row)
        self.db.flush()

        counts = {"working": 0, "not_working": 0, "review": 0}
        for source in parsed.rows:
            analysis = self._analyze_source(source, context)
            item = TechExpertActualizationItem(
                run_id=run.id,
                file_id=file_row.id,
                source_row=source.row_number,
                source_department=parsed.department_name,
                source_fio=source.fio,
                normalized_fio=normalize_fio(source.fio),
                source_position=source.position,
                source_email=source.email,
                source_phone=source.phone,
                source_login=source.login,
                source_password_encrypted=self.credential_box.encrypt(
                    source.password
                ),
                **analysis,
            )
            self.db.add(item)
            counts[str(analysis["category"])] += 1

        file_row.rows_count = len(parsed.rows)
        file_row.working_count = counts["working"]
        file_row.not_working_count = counts["not_working"]
        file_row.review_count = counts["review"]
        self.db.flush()
        self._recalculate_run(run)
        self.db.add(
            AuditLog(
                actor=actor,
                action="techexpert_actualization_file_analyzed",
                target=f"run:{run.id}",
                result="success",
                details=json.dumps(
                    {
                        "file_id": file_row.id,
                        "filename": file_row.filename,
                        "department": file_row.department_name,
                        **counts,
                    },
                    ensure_ascii=False,
                    sort_keys=True,
                ),
            )
        )
        self.db.commit()
        return {
            "file_id": file_row.id,
            "department": parsed.department_name,
            "total": len(parsed.rows),
            **counts,
        }

    def _recalculate_run(self, run: TechExpertActualizationRun) -> None:
        files = list(
            self.db.scalars(
                select(TechExpertActualizationFile).where(
                    TechExpertActualizationFile.run_id == run.id
                )
            ).all()
        )
        items = list(
            self.db.scalars(
                select(TechExpertActualizationItem).where(
                    TechExpertActualizationItem.run_id == run.id
                )
            ).all()
        )
        run.files_count = len(files)
        run.total_count = len(items)
        run.working_count = sum(item.category == "working" for item in items)
        run.not_working_count = sum(
            item.category == "not_working" for item in items
        )
        run.review_count = sum(item.category == "review" for item in items)
        run.ad_found_count = sum(
            item.ad_status in {"enabled", "disabled"} for item in items
        )
        run.updated_at = utcnow()

    def _recalculate_files(self, run_id: int) -> None:
        files = {
            file_row.id: file_row
            for file_row in self.db.scalars(
                select(TechExpertActualizationFile).where(
                    TechExpertActualizationFile.run_id == int(run_id)
                )
            ).all()
        }
        items_by_file: dict[int, list[TechExpertActualizationItem]] = {
            file_id: [] for file_id in files
        }
        for item in self.db.scalars(
            select(TechExpertActualizationItem).where(
                TechExpertActualizationItem.run_id == int(run_id)
            )
        ).all():
            items_by_file.setdefault(item.file_id, []).append(item)
        for file_id, file_row in files.items():
            items = items_by_file.get(file_id, [])
            file_row.rows_count = len(items)
            file_row.working_count = sum(
                item.category == "working" for item in items
            )
            file_row.not_working_count = sum(
                item.category == "not_working" for item in items
            )
            file_row.review_count = sum(
                item.category == "review" for item in items
            )

    def reanalyze_not_working(
        self,
        *,
        run_id: int,
        actor: str,
    ) -> dict[str, int]:
        run = self.get_run(run_id)
        if run.status != "open":
            raise ValueError("Пакет уже завершён")
        items = list(
            self.db.scalars(
                select(TechExpertActualizationItem)
                .where(
                    TechExpertActualizationItem.run_id == run.id,
                    TechExpertActualizationItem.category == "not_working",
                )
                .order_by(TechExpertActualizationItem.id)
            ).all()
        )
        result = {
            "checked": len(items),
            "working": 0,
            "not_working": 0,
            "review": 0,
        }
        if not items:
            return result

        context = self._analysis_context()
        for item in items:
            source = ParsedTechExpertRow(
                row_number=item.source_row,
                fio=item.source_fio,
                position=item.source_position,
                email=normalize_email(item.source_email),
                phone=item.source_phone,
                login=item.source_login,
                password="",
            )
            analysis = self._analyze_source(source, context)
            self._apply_analysis(item, analysis)
            result[str(analysis["category"])] += 1

        self.db.flush()
        self._recalculate_files(run.id)
        self._recalculate_run(run)
        self.db.add(
            AuditLog(
                actor=actor,
                action="techexpert_actualization_not_working_reanalyzed",
                target=f"run:{run.id}",
                result="success",
                details=json.dumps(result, ensure_ascii=False, sort_keys=True),
            )
        )
        self.db.commit()
        return result

    def run_details(self, run_id: int) -> dict[str, object]:
        run = self.get_run(run_id)
        files = list(
            self.db.scalars(
                select(TechExpertActualizationFile)
                .where(TechExpertActualizationFile.run_id == run.id)
                .order_by(TechExpertActualizationFile.id)
            ).all()
        )
        items = list(
            self.db.scalars(
                select(TechExpertActualizationItem)
                .where(TechExpertActualizationItem.run_id == run.id)
                .order_by(
                    TechExpertActualizationItem.category,
                    TechExpertActualizationItem.source_fio,
                    TechExpertActualizationItem.id,
                )
            ).all()
        )
        rows: list[dict[str, object]] = []
        for item in items:
            try:
                candidates = json.loads(item.candidates_json or "[]")
            except (TypeError, json.JSONDecodeError):
                candidates = []
            rows.append({"item": item, "candidates": candidates})
        return {"run": run, "files": files, "rows": rows}

    def resolve_item(
        self,
        *,
        run_id: int,
        item_id: int,
        record_id: int,
        ad_login: str,
        actor: str,
    ) -> None:
        run = self.get_run(run_id)
        if run.status != "open":
            raise ValueError("Пакет уже завершён")
        item = self.db.get(TechExpertActualizationItem, int(item_id))
        record = self.db.get(HRSourceRecord, int(record_id))
        if item is None or item.run_id != run.id:
            raise LookupError("Строка актуализации не найдена")
        if record is None or normalize_email(record.source_id) != self.source_id:
            raise ValueError("Выбранный работник относится к другой организации")

        mapping = HRRegistryManualMappingService(
            self.settings,
            self.db,
        ).save_ad_identifier(
            record_id=record.id,
            identifier=ad_login,
            actor=actor,
        )
        ad_user = ActiveDirectoryService(self.settings).get_user(
            str(mapping["ad_login"])
        )
        if ad_user is None:
            raise ValueError("AD-учетка не найдена после сопоставления")

        active = record.worker_key in self._active_worker_keys()
        snapshot = placement_snapshot(record)
        item.worker_key = record.worker_key
        item.hr_record_id = record.id
        item.current_fio = record.fio
        item.current_positions = json.dumps(snapshot["positions"], ensure_ascii=False)
        item.current_departments = json.dumps(
            snapshot["departments"], ensure_ascii=False
        )
        item.ad_login = ad_user.username
        item.ad_object_guid = ad_user.object_guid
        item.ad_distinguished_name = ad_user.distinguished_name
        item.ad_status = "enabled" if ad_user.is_enabled else "disabled"
        item.membership_state = (
            "member"
            if ActiveDirectoryService(self.settings).is_user_member_of_group(
                ad_user.username,
                self.config.ad_group_dn,
                object_guid=ad_user.object_guid,
            )
            else "not_member"
        )
        item.category = (
            "working" if active and ad_user.is_enabled
            else "not_working" if not active
            else "review"
        )
        item.reason = (
            "Сопоставление подтверждено оператором"
            if item.category != "review"
            else "Сопоставление подтверждено, но AD-учетка отключена"
        )
        self._recalculate_run(run)
        self.db.commit()

    def apply_group(self, *, run_id: int, action: str, actor: str) -> dict[str, int]:
        run = self.get_run(run_id)
        if run.status != "open":
            raise ValueError("Пакет уже завершён")
        if action not in {"add", "remove"}:
            raise ValueError("Неизвестное действие с группой")

        category = "working" if action == "add" else "not_working"
        items = list(
            self.db.scalars(
                select(TechExpertActualizationItem).where(
                    TechExpertActualizationItem.run_id == run.id,
                    TechExpertActualizationItem.category == category,
                    TechExpertActualizationItem.ad_object_guid != "",
                )
            ).all()
        )
        active_keys = self._active_worker_keys()
        unique: dict[str, TechExpertActualizationItem] = {}
        for item in items:
            unique.setdefault(item.ad_object_guid or item.ad_login, item)

        ad = ActiveDirectoryService(self.settings)
        result = {"selected": len(unique), "changed": 0, "skipped": 0, "errors": 0}
        for item in unique.values():
            record = self.db.get(HRSourceRecord, item.hr_record_id) if item.hr_record_id else None
            is_active = bool(record and record.worker_key in active_keys)
            if (action == "add" and not is_active) or (action == "remove" and is_active):
                item.group_action = "skipped"
                item.group_action_error = "Кадровое состояние изменилось перед действием"
                result["skipped"] += 1
                continue
            try:
                if action == "add":
                    state = ad.ensure_user_in_group(
                        item.ad_login,
                        self.config.ad_group_dn,
                        object_guid=item.ad_object_guid,
                    )
                    if state != "dry_run":
                        item.membership_state = "member"
                    if record is not None and state != "dry_run":
                        record.techexpert_access = True
                else:
                    state = ad.remove_user_from_group(
                        item.ad_login,
                        self.config.ad_group_dn,
                        object_guid=item.ad_object_guid,
                    )
                    if state != "dry_run":
                        item.membership_state = "not_member"
                    if record is not None and state != "dry_run":
                        record.techexpert_access = False
                item.group_action = state
                item.group_action_error = ""
                if state in {"added", "removed"}:
                    result["changed"] += 1
                else:
                    result["skipped"] += 1
            except Exception as exc:
                item.group_action = "failed"
                item.group_action_error = str(exc)[:4000]
                result["errors"] += 1

        self.db.add(
            AuditLog(
                actor=actor,
                action=f"techexpert_actualization_group_{action}",
                target=f"run:{run.id}",
                result="partial" if result["errors"] else "success",
                details=json.dumps(result, ensure_ascii=False, sort_keys=True),
            )
        )
        self.db.commit()
        return result

    def complete_run(self, *, run_id: int, actor: str) -> None:
        run = self.get_run(run_id)
        if run.status != "open":
            return
        run.status = "completed"
        run.completed_at = utcnow()
        run.updated_at = utcnow()
        self.db.add(
            AuditLog(
                actor=actor,
                action="techexpert_actualization_completed",
                target=f"run:{run.id}",
                result="success",
                details=json.dumps(
                    {
                        "files": run.files_count,
                        "total": run.total_count,
                        "working": run.working_count,
                        "not_working": run.not_working_count,
                        "review": run.review_count,
                    },
                    ensure_ascii=False,
                    sort_keys=True,
                ),
            )
        )
        self.db.commit()

    @staticmethod
    def _finish_workbook(workbook: Workbook) -> bytes:
        stream = io.BytesIO()
        workbook.save(stream)
        return stream.getvalue()

    def export_not_working(self, run_id: int) -> bytes:
        run = self.get_run(run_id)
        items = list(
            self.db.scalars(
                select(TechExpertActualizationItem)
                .where(
                    TechExpertActualizationItem.run_id == run.id,
                    TechExpertActualizationItem.category == "not_working",
                )
                .order_by(
                    TechExpertActualizationItem.source_department,
                    TechExpertActualizationItem.source_fio,
                    TechExpertActualizationItem.id,
                )
            ).all()
        )
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Не работают"
        row_number = 1
        current_department = None
        sequence = 0
        for item in items:
            if item.source_department != current_department:
                if row_number > 1:
                    row_number += 1
                current_department = item.source_department
                sequence = 0
                sheet.cell(row_number, 1, safe_excel_value(current_department))
                sheet.cell(row_number, 1).font = Font(bold=True, size=12)
                row_number += 1
                headers = ["№ п/п", "ФИО", "Должность", "E-mail", "Телефон"]
                for column, header in enumerate(headers, 1):
                    cell = sheet.cell(row_number, column, header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill("solid", fgColor="DCE6F1")
                row_number += 1
            sequence += 1
            values = [
                sequence,
                item.source_fio,
                item.source_position,
                item.source_email,
                item.source_phone,
            ]
            for column, value in enumerate(values, 1):
                sheet.cell(
                    row_number,
                    column,
                    value if isinstance(value, int) else safe_excel_value(value),
                )
            row_number += 1
        for column, width in {"A": 10, "B": 42, "C": 44, "D": 34, "E": 22}.items():
            sheet.column_dimensions[column].width = width
        return self._finish_workbook(workbook)

    def export_current(self) -> bytes:
        records = list(
            self.db.scalars(
                select(HRSourceRecord)
                .where(
                    HRSourceRecord.source_id == self.source_id,
                    HRSourceRecord.techexpert_access.is_(True),
                )
                .order_by(HRSourceRecord.fio)
            ).all()
        )
        worker_keys = {record.worker_key for record in records}
        source_fields: dict[str, dict[str, str]] = {}
        if worker_keys:
            source_items = list(
                self.db.scalars(
                    select(TechExpertActualizationItem)
                    .where(
                        TechExpertActualizationItem.worker_key.in_(worker_keys)
                    )
                    .order_by(TechExpertActualizationItem.id.desc())
                ).all()
            )
            for item in source_items:
                values = source_fields.setdefault(
                    item.worker_key,
                    {"phone": "", "login": "", "password": ""},
                )
                if not values["phone"] and item.source_phone:
                    values["phone"] = item.source_phone
                if not values["login"] and item.source_login:
                    values["login"] = item.source_login
                if (
                    not values["password"]
                    and item.source_password_encrypted
                ):
                    values["password"] = self.credential_box.decrypt(
                        item.source_password_encrypted
                    )

        grouped: dict[str, list[tuple[HRSourceRecord, str]]] = {}
        for record in records:
            snapshot = placement_snapshot(record)
            placements = []
            try:
                raw = json.loads(record.placements_json or "[]")
            except (TypeError, json.JSONDecodeError):
                raw = []
            for value in raw if isinstance(raw, list) else []:
                if not isinstance(value, dict):
                    continue
                department = normalize_text(value.get("department"))
                position = normalize_text(value.get("position"))
                top = department.split(" / ", 1)[0].strip() or "Без подразделения"
                placements.append((top, position))
            if not placements:
                tops = snapshot["top_departments"] or ["Без подразделения"]
                positions = snapshot["positions"] or [""]
                placements = [(tops[0], position) for position in positions]
            for top, position in placements:
                grouped.setdefault(top, []).append((record, position))

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Актуальный список"
        row_number = 1
        for department in sorted(grouped, key=str.casefold):
            if row_number > 1:
                row_number += 1
            sheet.cell(row_number, 1, safe_excel_value(department))
            sheet.cell(row_number, 1).font = Font(bold=True, size=12)
            row_number += 1
            for column, header in enumerate(
                [
                    "№ п/п",
                    "ФИО",
                    "Должность",
                    "E-mail",
                    "Телефон",
                    "Логин",
                    "Пароль",
                ],
                1,
            ):
                cell = sheet.cell(row_number, column, header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="DCE6F1")
            row_number += 1
            for sequence, (record, position) in enumerate(
                sorted(grouped[department], key=lambda item: item[0].fio.casefold()),
                1,
            ):
                values = [
                    sequence,
                    record.fio,
                    position,
                    record.corporate_email,
                    source_fields.get(record.worker_key, {}).get("phone", ""),
                    source_fields.get(record.worker_key, {}).get("login", ""),
                    source_fields.get(record.worker_key, {}).get("password", ""),
                ]
                for column, value in enumerate(values, 1):
                    if column == 7:
                        excel_value = safe_excel_secret_value(value)
                    else:
                        excel_value = (
                            value
                            if isinstance(value, int)
                            else safe_excel_value(value)
                        )
                    sheet.cell(
                        row_number,
                        column,
                        excel_value,
                    )
                row_number += 1
        for column, width in {
            "A": 10,
            "B": 42,
            "C": 44,
            "D": 34,
            "E": 22,
            "F": 24,
            "G": 24,
        }.items():
            sheet.column_dimensions[column].width = width
        return self._finish_workbook(workbook)

    def access_summary(self) -> dict[str, int]:
        return {
            "access_count": int(
                self.db.scalar(
                    select(func.count(HRSourceRecord.id)).where(
                        HRSourceRecord.source_id == self.source_id,
                        HRSourceRecord.techexpert_access.is_(True),
                    )
                )
                or 0
            )
        }
